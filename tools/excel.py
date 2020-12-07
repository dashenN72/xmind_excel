# -*- coding: utf-8 -*-
# @Time    : 2020/2/4 0:31
# @Author  : dashenN72

"""
excel操作
"""

import string
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import config


class ExcelWriter(object):
    def __init__(self, excel_name, module_excel):
        self.wb = openpyxl.Workbook()
        self.sheet = self.wb.active
        self.sheet.title = module_excel[0][0]
        self.excel_name = excel_name
        self.sheets = module_excel[1:]

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.wb.save(r".\\output\\" + self.excel_name)

    def __get_data_by_column_name(self, column_name):
        first_column = self.sheet[column_name]
        data_column = [first_column[x].value for x in range(1, len(first_column))]
        return data_column

    def __get_maxlength(self, column_name):
        """
        获取一个类型为object的Series中的最大占位长度，用于确定导出的xlsx文件的列宽
        :param column_name: 列编号(A,B,C)
        :return: 列中最大的长度值
        """
        str_list = self.__get_data_by_column_name(column_name)
        len_list = []
        for elem in str_list:
            length = 0
            if elem is not None:  # 单元格值非空
                elem_split = list(elem)  # 字符串拆分成独立字符组成的list
                for c in elem_split:
                    if ord(c) <= 256:
                        length += 1
                    else:
                        length += 2
            else:
                pass
            len_list.append(length)
        return max(len_list)

    def __auto_width(self):
        """
        自动设置单元格宽度
        :return: None
        """
        cols_list = [param[0] for param in self.sheets]  # 获取列名
        for i in range(0, len(cols_list)):
            letter = chr(i + 65)  # 由ASCII值获得对应的列字母
            max_len = self.__get_maxlength(letter)
            if max_len <= 10:
                self.sheet.column_dimensions[letter].width = 10
            elif max_len <= 50:
                self.sheet.column_dimensions[letter].width = max_len + 1
            else:
                self.sheet.column_dimensions[letter].width = 50
                for cell in self.sheet[letter]:
                    cell.alignment = Alignment(wrap_text=True)

    def init_title(self):
        """
        初始化excel中sheet的标题
        :return:
        """
        sheets = self.sheets  # 获取sheet数据
        names_cols = [param + '1' for param in list(string.ascii_uppercase)[:len(sheets)]]  # 生成excel表列编号
        for name in names_cols:
            self.sheet[name] = sheets[names_cols.index(name)][0]
        # 给用例标题加粗
        for col in range(1, len(sheets)+1):
            self.sheet.cell(row=1, column=col).font = Font(bold=True)  # 文字加粗
            self.sheet.cell(row=1, column=col).alignment = Alignment(horizontal='center', vertical='center')  # 居中

    def write_rows(self, sheet_data):
        """
        写数据到excel
        :param sheet_data: list形式的数据
        :return:
        """
        num_had_write = 0
        for row in range(1, len(sheet_data)+1):
            for col in range(1, len(sheet_data[row-1])+1):
                self.sheet.cell(row=row+1, column=col).value = str(sheet_data[row-1][col-1])
                num_had_write += 1
                # self.sheet.cell(row=row, column=8).font = Font(color=RED)
        self.__auto_width()  # 自动设置单元格宽度
        if num_had_write == len(sheet_data):  # 所有数据写入excel成功
            return True
        else:
            return False


if __name__ == "__main__":
    pass
